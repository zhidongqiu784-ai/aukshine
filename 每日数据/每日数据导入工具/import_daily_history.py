#!/usr/bin/env python3
r"""Validate and upsert historical daily data from the approved Excel template.

The default mode is a read-only dry run. Production writes require both
--execute and --confirm-production WRITE-PRODUCTION.

python "D:\wayne\每日数据\每日数据导入工具\import_daily_history.py" --file "<Excel文件路径>" --validate-only
python "D:\wayne\每日数据\每日数据导入工具\import_daily_history.py" --file "<Excel文件路径>" --timeout 120 --execute --confirm-production WRITE-PRODUCTION
python "D:\wayne\每日数据\每日数据导入工具\import_daily_history.py" --folder "C:\Users\Aniya\Desktop\每日数据历史导入" --validate-only
python "D:\wayne\每日数据\每日数据导入工具\import_daily_history.py" --folder "C:\Users\Aniya\Desktop\每日数据历史导入" --timeout 120 --execute --confirm-production WRITE-PRODUCTION
"""

from __future__ import annotations

import argparse
import json
import os
import posixpath
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
except ImportError as exc:  # pragma: no cover - environment-dependent message
    raise SystemExit("缺少 openpyxl，请先运行：python -m pip install openpyxl") from exc


DEFAULT_API_BASE_URL = "https://erp.aukshine.com/api"
DEFAULT_EXCEL_FILE = Path(__file__).with_name("导入历史数据模版.xlsx")
LOCAL_CONFIG_FILE = Path(__file__).with_name("import_daily_history.local.json")


def load_local_config() -> dict[str, Any]:
    if not LOCAL_CONFIG_FILE.exists():
        return {}
    try:
        config = json.loads(LOCAL_CONFIG_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit(f"本地配置读取失败：{LOCAL_CONFIG_FILE}：{exc}") from exc
    if not isinstance(config, dict):
        raise SystemExit(f"本地配置必须是 JSON 对象：{LOCAL_CONFIG_FILE}")
    return config


LOCAL_CONFIG = load_local_config()
DEFAULT_API_TOKEN = str(LOCAL_CONFIG.get("api_token") or "").strip()
DEFAULT_SHEET = "导入数据"
SHEET_ALIASES = ("导入数据", "数据导入")
EXECUTE_CONFIRMATION = "WRITE-PRODUCTION"
KEYWORD_HEADER_RE = re.compile(r"^词\s*\d+\s*[:：]\s*(.+?)\s*自然位\s*$", re.IGNORECASE)
COMPETITOR_HEADER_RE = re.compile(r"^竞对\s*([1-9]\d*)\s*(ASIN|排名|操作分析)\s*$", re.IGNORECASE)
SCREENSHOT_FIELD_NAMES = {
    "page_screenshot",
    "review_screenshot",
    "bad_review_notes",
    "keyword_trend_screenshot",
    "ad_framework_screenshot",
    "keyword_performance_screenshot",
    "link_problem",
    "operation_record",
    "review_notes",
    "ad_optimization_logs",
}
IMAGE_CONTENT_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
}
CONFIG_LABELS = {
    "country": ("国家", "国家（下拉选择）"),
    "asin": ("ASIN",),
    "start_date": ("导入数据起始日期",),
    "end_date": ("导入数据终止日期",),
    "competitor1_asin": ("竞对1 ASIN",),
    "competitor2_asin": ("竞对2 ASIN",),
    "competitor3_asin": ("竞对3 ASIN",),
}


@dataclass(frozen=True)
class FieldSpec:
    resource: str
    field: str
    value_type: str = "text"


STATIC_FIELDS: dict[str, FieldSpec] = {
    "推广天数": FieldSpec("daily_asins", "promotion_days", "integer"),
    "活动标注": FieldSpec("daily_asins", "activity_annotation"),
    "LP/WP/TP": FieldSpec("daily_asins", "list_price", "decimal"),
    "购物车价格": FieldSpec("daily_asins", "daily_price", "decimal"),
    "折后售价": FieldSpec("daily_asins", "price_after_discount", "decimal"),
    "售卖账号": FieldSpec("daily_asins", "selling_accounts"),
    "目标拆解-小类排名": FieldSpec("target_management", "target_subcategory_rank", "integer"),
    "目标拆解 - 小类排名": FieldSpec("target_management", "target_subcategory_rank", "integer"),
    "目标拆解-单量": FieldSpec("target_management", "target_order_qty", "integer"),
    "目标拆解 - 单量": FieldSpec("target_management", "target_order_qty", "integer"),
    "实际总单量": FieldSpec("weekly_performance", "sales", "integer"),
    "目标差距": FieldSpec("daily_asins", "target_gap", "integer"),
    "测评折后价": FieldSpec("daily_order_link_tracking", "review_discounted_price", "decimal"),
    "①测评单": FieldSpec("daily_asins", "rsg_number", "integer"),
    "手机端流量": FieldSpec("weekly_performance", "sessions_mobile", "integer"),
    "电脑端流量": FieldSpec("weekly_performance", "sessions", "integer"),
    "汇总流量-会话量": FieldSpec("weekly_performance", "zongliuliang", "integer"),
    "页面浏览量": FieldSpec("weekly_performance", "page_views_total", "integer"),
    "广告点击": FieldSpec("weekly_performance", "guanggaodianji", "integer"),
    "退货量": FieldSpec("weekly_performance", "return_goods_count", "integer"),
    "退货率": FieldSpec("weekly_performance", "return_goods_rate", "ratio"),
    "小类排名": FieldSpec("weekly_performance", "ranking", "integer"),
    "自己页面截图": FieldSpec("daily_order_link_tracking", "page_screenshot"),
    "review数量": FieldSpec("daily_asins", "number_of_comments", "integer"),
    "星级": FieldSpec("daily_asins", "star_rating", "decimal"),
    "review详细截图": FieldSpec("daily_order_link_tracking", "review_screenshot"),
    "review 详细截图": FieldSpec("daily_order_link_tracking", "review_screenshot"),
    "差评rating / 差评": FieldSpec("daily_order_link_tracking", "bad_review_notes"),
    "差评 rating/差评": FieldSpec("daily_order_link_tracking", "bad_review_notes"),
    "Asin 西柚/sif 搜索词排名趋势截图": FieldSpec("daily_order_link_tracking", "keyword_trend_screenshot"),
    "Asin 广告框架截图": FieldSpec("daily_order_link_tracking", "ad_framework_screenshot"),
    "Asin 搜索词表现截图": FieldSpec("daily_order_link_tracking", "keyword_performance_screenshot"),
    "链接问题": FieldSpec("daily_order_link_tracking", "link_problem"),
    "今日操作记录": FieldSpec("daily_order_link_tracking", "operation_record"),
    "复盘": FieldSpec("daily_order_link_tracking", "review_notes"),
    "广告优化操作动作记录（大方向记录）": FieldSpec("daily_order_link_tracking", "ad_optimization_logs"),
    "广告优化操作动作记录 (大方向记录)": FieldSpec("daily_order_link_tracking", "ad_optimization_logs"),
    "曝光量": FieldSpec("weekly_performance", "impressions", "integer"),
    "广告曝光量": FieldSpec("weekly_performance", "impressions", "integer"),
    "广告点击量": FieldSpec("weekly_performance", "guanggaodianji", "integer"),
    "广告花费": FieldSpec("weekly_performance", "guanggaohuafei", "decimal"),
    "广告总单量": FieldSpec("weekly_performance", "guanggaodan", "integer"),
    "广告销售额": FieldSpec("weekly_performance", "ad_sales_amount", "decimal"),
    "间接跑单订单量": FieldSpec("weekly_performance", "indirect_order_volume", "integer"),
    "SP广告费": FieldSpec("weekly_performance", "ads_sp_cost", "decimal"),
    "SP 广告费": FieldSpec("weekly_performance", "ads_sp_cost", "decimal"),
    "SD广告费": FieldSpec("weekly_performance", "ads_sd_cost", "decimal"),
    "SD 广告费": FieldSpec("weekly_performance", "ads_sd_cost", "decimal"),
    "SB广告费": FieldSpec("weekly_performance", "shared_ads_sb_cost", "decimal"),
    "SB 广告费": FieldSpec("weekly_performance", "shared_ads_sb_cost", "decimal"),
    "SBV广告费": FieldSpec("weekly_performance", "shared_ads_sbv_cost", "decimal"),
    "SBV 广告费": FieldSpec("weekly_performance", "shared_ads_sbv_cost", "decimal"),
    "秒杀价格（当地币）": FieldSpec("daily_profit", "flash_sale_price", "decimal"),
    "秒杀总单量": FieldSpec("daily_profit", "flash_sale_qty", "integer"),
    "秒杀天数": FieldSpec("daily_profit", "flash_sale_days", "integer"),
}


RESOURCE_KEY_FIELDS = {
    "daily_asins": "country_asin_date",
    "target_management": "country_asin_date",
    "weekly_performance": "country_asin_week",
    "daily_order_link_tracking": "country_asin_date",
    "daily_profit": "country_asin_date",
}


@dataclass
class EmbeddedImage:
    row: int
    column: int
    filename: str
    content_type: str
    data: bytes


@dataclass
class RowData:
    excel_row: int
    country: str
    asin: str
    day: date
    country_asin: str
    asin_country: str
    country_asin_date: str
    country_asin_week: str
    static: dict[str, dict[str, Any]] = field(default_factory=dict)
    keywords: list[tuple[str, str]] = field(default_factory=list)
    keyword_clears: list[str] = field(default_factory=list)
    competitors: dict[str, dict[str, Any]] = field(default_factory=dict)
    images: dict[str, list[EmbeddedImage]] = field(default_factory=dict)
    competitor_images: dict[str, list[EmbeddedImage]] = field(default_factory=dict)


class ValidationError(Exception):
    pass


class ApiError(RuntimeError):
    pass


class ConnectionApiError(ApiError):
    pass


NUMERIC_VALUE_TYPES = frozenset({"integer", "decimal", "percent", "ratio"})
NUMERIC_BLANK_MARKERS = frozenset({"-", "/", "未找到", "#DIV/0!", "#VALUE!"})
KEYWORD_CLEAR_MARKERS = frozenset({"/"})
FORMULA_ERROR_MARKERS = frozenset(
    {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A"}
)
TEXT_FIELD_MAX_LENGTHS = {("daily_asins", "activity_annotation"): 255}
DISPIMG_RE = re.compile(
    r'^\s*=(?:_xlfn\.)?DISPIMG\(\s*"([^"]+)"\s*[,;]\s*\d+\s*\)\s*$',
    re.IGNORECASE,
)


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def is_numeric_placeholder(value: Any) -> bool:
    return isinstance(value, str) and value.strip() in NUMERIC_BLANK_MARKERS


def is_keyword_clear_marker(value: Any) -> bool:
    return isinstance(value, str) and value.strip() in KEYWORD_CLEAR_MARKERS


def dispimg_id(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    match = DISPIMG_RE.match(value)
    return match.group(1) if match else None


def is_dispimg_formula(value: Any) -> bool:
    return dispimg_id(value) is not None


def is_formula_error(value: Any) -> bool:
    return isinstance(value, str) and value.strip().upper() in FORMULA_ERROR_MARKERS


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_country(value: Any) -> str:
    return str(value or "").strip().upper()


def normalize_asin(value: Any) -> str:
    return str(value or "").strip().upper()


def normalize_name(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def parse_date(value: Any, epoch: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        parsed = from_excel(value, epoch=epoch)
        return parsed.date() if isinstance(parsed, datetime) else parsed
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError("日期必须是 Excel 日期或 YYYY-MM-DD")


def parse_decimal(value: Any) -> int | float:
    try:
        number = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        raise ValueError("必须是数字") from None
    if not number.is_finite():
        raise ValueError("必须是有限数字")
    return int(number) if number == number.to_integral_value() else float(number)


def parse_integer(value: Any) -> int:
    number = parse_decimal(value)
    if isinstance(number, float) and not number.is_integer():
        raise ValueError("必须是整数")
    return int(number)


def parse_percent(value: Any) -> float:
    text = str(value).strip()
    if text.endswith("%"):
        result = float(Decimal(text[:-1].replace(",", "").strip()) / Decimal(100))
    else:
        result = float(parse_decimal(value))
    if result < -1 or result > 1:
        raise ValueError("百分比必须在 -100% 到 100% 之间；请用 10% 而不是 10")
    return result


def parse_ratio(value: Any) -> float:
    text = str(value).strip()
    if text.endswith("%"):
        return float(Decimal(str(parse_decimal(text[:-1]))) / Decimal(100))
    return float(parse_decimal(value))


def convert_value(value: Any, value_type: str) -> Any:
    if value_type == "integer":
        return parse_integer(value)
    if value_type == "decimal":
        return parse_decimal(value)
    if value_type == "percent":
        return parse_percent(value)
    if value_type == "ratio":
        return parse_ratio(value)
    return str(value).strip() if not isinstance(value, str) else value.strip()


def decimal_number(value: Any) -> Decimal:
    return Decimal(str(value).replace(",", "").strip())


def decimal_output(value: Decimal) -> int | float:
    return int(value) if value == value.to_integral_value() else float(value)


def rounded_division(
    numerator: Any, denominator: Any, places: int, *, absolute: bool = False
) -> float | None:
    numerator_decimal = decimal_number(numerator)
    denominator_decimal = decimal_number(denominator)
    if denominator_decimal == 0:
        return None
    result = numerator_decimal / denominator_decimal
    if absolute:
        result = abs(result)
    quantum = Decimal(1).scaleb(-places)
    return float(result.quantize(quantum, rounding=ROUND_HALF_UP))


def calculate_weekly_derived_fields(fields: dict[str, Any]) -> dict[str, Any]:
    calculated: dict[str, Any] = {}

    def has_all(*names: str) -> bool:
        return all(name in fields and not is_blank(fields[name]) for name in names)

    if has_all("sales", "guanggaodan"):
        calculated["zirandan"] = decimal_output(
            decimal_number(fields["sales"]) - decimal_number(fields["guanggaodan"])
        )
        calculated["adv_rate"] = rounded_division(fields["guanggaodan"], fields["sales"], 4)

    if has_all("zongliuliang", "guanggaodianji"):
        organic_traffic = decimal_number(fields["zongliuliang"]) - decimal_number(fields["guanggaodianji"])
        organic_traffic_output = decimal_output(organic_traffic)
        calculated["organic_traffic"] = organic_traffic_output
        calculated["zirandianji"] = organic_traffic_output
        calculated["natural_traffic_proportion"] = rounded_division(
            organic_traffic, fields["zongliuliang"], 4
        )

    if has_all("guanggaodan", "guanggaodianji"):
        calculated["guanggaocvr"] = rounded_division(
            fields["guanggaodan"], fields["guanggaodianji"], 4
        )

    if has_all("guanggaodianji", "impressions"):
        calculated["ctr"] = rounded_division(fields["guanggaodianji"], fields["impressions"], 4)

    if has_all("guanggaohuafei", "guanggaodianji"):
        calculated["cpc"] = rounded_division(fields["guanggaohuafei"], fields["guanggaodianji"], 2)

    if has_all("guanggaohuafei", "ad_sales_amount"):
        calculated["acos"] = rounded_division(fields["guanggaohuafei"], fields["ad_sales_amount"], 4)

    if has_all("guanggaohuafei", "guanggaodan"):
        calculated["cpa"] = rounded_division(fields["guanggaohuafei"], fields["guanggaodan"], 2)
        calculated["cpo"] = rounded_division(
            fields["guanggaohuafei"], fields["guanggaodan"], 2, absolute=True
        )

    if has_all("guanggaohuafei", "sales"):
        calculated["cpu"] = rounded_division(
            fields["guanggaohuafei"], fields["sales"], 2, absolute=True
        )

    if has_all("sales", "zongliuliang"):
        conversion_rate = rounded_division(fields["sales"], fields["zongliuliang"], 4)
        calculated["volume_cvr"] = conversion_rate
        calculated["session_conversion_rate"] = conversion_rate

    return calculated


def build_keys(country: str, asin: str, day: date) -> tuple[str, str, str, str]:
    country_asin = f"{country}_{asin}"
    asin_country = f"{asin}_{country}"
    dated_key = f"{country_asin}_{day.isoformat()}"
    return country_asin, asin_country, dated_key, dated_key


def values_equivalent(left: Any, right: Any) -> bool:
    if left == right:
        return True
    try:
        return Decimal(str(left)) == Decimal(str(right))
    except (InvalidOperation, ValueError):
        return str(left).strip() == str(right).strip()


def find_header_row(ws: Any) -> int:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        headers = {clean_header(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)}
        if "日期" in headers:
            return row_idx
    raise ValidationError("找不到包含“日期”的表头行")


def find_config_value(ws: Any, header_row: int, labels: Iterable[str]) -> Any:
    accepted_labels = set(labels)
    for row_idx in range(1, header_row):
        for col_idx in range(1, ws.max_column + 1):
            if clean_header(ws.cell(row_idx, col_idx).value) in accepted_labels:
                return ws.cell(row_idx, col_idx + 1).value
    return None


def load_template_keywords(workbook: Any, country: str, asin: str) -> list[str]:
    if not country or not asin or "模板配置" not in workbook.sheetnames:
        return []
    ws = workbook["模板配置"]
    header_columns = {
        clean_header(ws.cell(1, col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
        if clean_header(ws.cell(1, col_idx).value)
    }
    country_col = header_columns.get("国家")
    asin_col = header_columns.get("ASIN")
    keyword_col = header_columns.get("关键词")
    order_col = header_columns.get("顺序")
    if not country_col or not keyword_col:
        return []
    matches: list[tuple[int, str]] = []
    seen: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        row_country = normalize_country(ws.cell(row_idx, country_col).value)
        row_asin = normalize_asin(ws.cell(row_idx, asin_col).value) if asin_col else ""
        keyword = str(ws.cell(row_idx, keyword_col).value or "").strip()
        if row_country != country or (asin_col and row_asin != asin) or not keyword:
            continue
        normalized = normalize_name(keyword)
        if normalized in seen:
            continue
        seen.add(normalized)
        try:
            raw_order = ws.cell(row_idx, order_col).value if order_col else None
            order = int(raw_order) if not is_blank(raw_order) else len(matches) + 1
        except (TypeError, ValueError):
            order = len(matches) + 1
        matches.append((order, keyword))
    return [keyword for _order, keyword in sorted(matches, key=lambda item: item[0])]


def resolve_xlsx_part(base_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), target))


def relationships_path(part: str) -> str:
    return posixpath.join(posixpath.dirname(part), "_rels", posixpath.basename(part) + ".rels")


def parse_relationships(archive: zipfile.ZipFile, part: str) -> dict[str, str]:
    rel_path = relationships_path(part)
    if rel_path not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read(rel_path))
    result: dict[str, str] = {}
    for relation in root:
        if relation.get("TargetMode") == "External":
            continue
        relation_id = relation.get("Id")
        target = relation.get("Target")
        if relation_id and target:
            result[relation_id] = resolve_xlsx_part(part, target)
    return result


def xlsx_content_types(archive: zipfile.ZipFile) -> tuple[dict[str, str], dict[str, str]]:
    root = ET.fromstring(archive.read("[Content_Types].xml"))
    defaults: dict[str, str] = {}
    overrides: dict[str, str] = {}
    for item in root:
        extension = item.get("Extension")
        part_name = item.get("PartName")
        content_type = item.get("ContentType")
        if extension and content_type:
            defaults[extension.lower()] = content_type
        elif part_name and content_type:
            overrides[part_name.lstrip("/")] = content_type
    return defaults, overrides


def excel_column_name(column: int) -> str:
    result = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        result = chr(65 + remainder) + result
    return result


def excel_column_number(cell_reference: str) -> int:
    match = re.fullmatch(r"([A-Z]+)\d+", cell_reference.upper())
    if not match:
        raise ValueError(f"Excel 单元格坐标无效：{cell_reference}")
    result = 0
    for character in match.group(1):
        result = result * 26 + ord(character) - 64
    return result


def extract_embedded_images(path: Path, sheet_name: str) -> dict[tuple[int, int], list[EmbeddedImage]]:
    spreadsheet_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    drawing_ns = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_main_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
    wps_cell_image_ns = "http://www.wps.cn/officeDocument/2017/etCustomData"
    namespaces = {"xdr": drawing_ns, "a": drawing_main_ns, "r": office_rel_ns}

    try:
        with zipfile.ZipFile(path) as archive:
            workbook_part = "xl/workbook.xml"
            workbook_root = ET.fromstring(archive.read(workbook_part))
            relation_id = None
            for sheet in workbook_root.findall(f".//{{{spreadsheet_ns}}}sheet"):
                if sheet.get("name") == sheet_name:
                    relation_id = sheet.get(f"{{{office_rel_ns}}}id")
                    break
            if not relation_id:
                return {}
            sheet_part = parse_relationships(archive, workbook_part).get(relation_id)
            if not sheet_part:
                return {}
            sheet_root = ET.fromstring(archive.read(sheet_part))
            default_types, override_types = xlsx_content_types(archive)
            result: dict[tuple[int, int], list[EmbeddedImage]] = {}
            image_number = 0

            drawing_node = sheet_root.find(f"{{{spreadsheet_ns}}}drawing")
            if drawing_node is not None:
                drawing_id = drawing_node.get(f"{{{office_rel_ns}}}id")
                drawing_part = parse_relationships(archive, sheet_part).get(str(drawing_id))
                if drawing_part:
                    drawing_root = ET.fromstring(archive.read(drawing_part))
                    drawing_relations = parse_relationships(archive, drawing_part)
                    for anchor_name in ("oneCellAnchor", "twoCellAnchor"):
                        for anchor in drawing_root.findall(f"xdr:{anchor_name}", namespaces):
                            start = anchor.find("xdr:from", namespaces)
                            blip = anchor.find(".//a:blip", namespaces)
                            if start is None or blip is None:
                                continue
                            row_node = start.find("xdr:row", namespaces)
                            column_node = start.find("xdr:col", namespaces)
                            media_id = blip.get(f"{{{office_rel_ns}}}embed")
                            if row_node is None or column_node is None or not media_id:
                                continue
                            media_part = drawing_relations.get(media_id)
                            if not media_part or media_part not in archive.namelist():
                                raise ValidationError(f"Excel 图片引用损坏：{media_id}")
                            row = int(row_node.text or 0) + 1
                            column = int(column_node.text or 0) + 1
                            extension = Path(media_part).suffix.lower()
                            content_type = override_types.get(media_part) or default_types.get(extension.lstrip("."))
                            content_type = content_type or IMAGE_CONTENT_TYPES.get(extension, "application/octet-stream")
                            image_number += 1
                            filename = f"screenshot-r{row}-c{column}-{image_number}{extension or '.bin'}"
                            image = EmbeddedImage(row, column, filename, content_type, archive.read(media_part))
                            result.setdefault((row, column), []).append(image)

            cell_images_part = "xl/cellimages.xml"
            if cell_images_part in archive.namelist():
                cell_images_root = ET.fromstring(archive.read(cell_images_part))
                cell_image_relations = parse_relationships(archive, cell_images_part)
                cell_image_media: dict[str, str] = {}
                for cell_image in cell_images_root.findall(f"{{{wps_cell_image_ns}}}cellImage"):
                    properties = cell_image.find(".//xdr:cNvPr", namespaces)
                    blip = cell_image.find(".//a:blip", namespaces)
                    if properties is None or blip is None:
                        continue
                    image_id = properties.get("name")
                    relation_id = blip.get(f"{{{office_rel_ns}}}embed")
                    media_part = cell_image_relations.get(str(relation_id))
                    if image_id and media_part:
                        cell_image_media[image_id] = media_part

                for cell in sheet_root.findall(f".//{{{spreadsheet_ns}}}c"):
                    formula = cell.find(f"{{{spreadsheet_ns}}}f")
                    if formula is None or not formula.text:
                        continue
                    image_id = dispimg_id(f"={formula.text}")
                    if not image_id:
                        continue
                    coordinate = cell.get("r") or ""
                    media_part = cell_image_media.get(image_id)
                    if not media_part or media_part not in archive.namelist():
                        raise ValidationError(f"WPS 单元格图片引用损坏：{coordinate} / {image_id}")
                    row_match = re.search(r"\d+$", coordinate)
                    if not row_match:
                        raise ValidationError(f"WPS 单元格图片坐标无效：{coordinate}")
                    row = int(row_match.group())
                    column = excel_column_number(coordinate)
                    extension = Path(media_part).suffix.lower()
                    content_type = override_types.get(media_part) or default_types.get(extension.lstrip("."))
                    content_type = content_type or IMAGE_CONTENT_TYPES.get(extension, "application/octet-stream")
                    image_number += 1
                    filename = f"screenshot-r{row}-c{column}-{image_number}{extension or '.bin'}"
                    image = EmbeddedImage(row, column, filename, content_type, archive.read(media_part))
                    result.setdefault((row, column), []).append(image)
            return result
    except (zipfile.BadZipFile, KeyError, ET.ParseError, ValueError) as exc:
        raise ValidationError(f"无法读取 Excel 嵌入图片：{exc}") from exc


def read_rows(path: Path, sheet_name: str) -> tuple[list[RowData], list[str]]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    if sheet_name not in workbook.sheetnames:
        if sheet_name == DEFAULT_SHEET:
            sheet_name = next((name for name in SHEET_ALIASES if name in workbook.sheetnames), sheet_name)
        if sheet_name not in workbook.sheetnames:
            raise ValidationError(f"工作表不存在：{sheet_name}；现有工作表：{', '.join(workbook.sheetnames)}")
    ws = workbook[sheet_name]
    values_workbook = load_workbook(path, data_only=True, read_only=True)
    values_ws = values_workbook[sheet_name]
    header_row = find_header_row(ws)
    config = {
        name: find_config_value(ws, header_row, label)
        for name, label in CONFIG_LABELS.items()
    }
    country = normalize_country(config["country"])
    asin = normalize_asin(config["asin"])
    template_keywords = load_template_keywords(workbook, country, asin)
    errors: list[str] = []
    top_competitor_asins = {
        f"竞对{index}": normalize_asin(config[f"competitor{index}_asin"])
        for index in range(1, 4)
    }
    seen_top_competitor_asins: dict[str, str] = {}
    for role, competitor_asin in top_competitor_asins.items():
        if not competitor_asin:
            continue
        previous_role = seen_top_competitor_asins.get(competitor_asin)
        if previous_role:
            errors.append(f"顶部配置：{previous_role}和{role}重复填写竞对 ASIN {competitor_asin}")
        else:
            seen_top_competitor_asins[competitor_asin] = role
    headers: dict[int, str] = {}
    seen_headers: dict[str, int] = {}
    dynamic_keyword_slots: dict[int, int] = {}

    for col in range(1, ws.max_column + 1):
        raw_header = ws.cell(header_row, col).value
        header = clean_header(raw_header)
        if isinstance(raw_header, str) and raw_header.lstrip().startswith("=") and "模板配置" in raw_header:
            slot_match = re.search(r'"词\s*(\d+)\s*[:：]', raw_header, re.IGNORECASE)
            if not slot_match:
                errors.append(f"第 {col} 列动态关键词公式无法识别")
                continue
            slot = int(slot_match.group(1))
            dynamic_keyword_slots[col] = slot
            if slot <= len(template_keywords):
                header = f"词{slot}:{template_keywords[slot - 1]} 自然位"
            else:
                continue
        if not header:
            continue
        if header in seen_headers:
            errors.append(f"表头重复：{header}（第 {seen_headers[header]} 列和第 {col} 列）")
        headers[col] = header
        seen_headers[header] = col

    required = {"日期"}
    missing = sorted(required - set(seen_headers))
    if missing:
        errors.append(f"缺少必填列：{', '.join(missing)}")

    keyword_area_columns: set[int] = set()
    keyword_area_start = seen_headers.get("Asin 搜索词表现截图")
    keyword_area_end = seen_headers.get("链接问题")
    if keyword_area_start and keyword_area_end and keyword_area_start < keyword_area_end:
        keyword_area_columns.update(range(keyword_area_start + 1, keyword_area_end))

    recognized = required | set(STATIC_FIELDS)
    unknown_headers = []
    keyword_columns: dict[int, str] = {}
    competitor_columns: dict[int, tuple[str, str]] = {}
    for col, header in headers.items():
        keyword_match = KEYWORD_HEADER_RE.match(header)
        competitor_match = COMPETITOR_HEADER_RE.match(header)
        if keyword_match:
            keyword_columns[col] = keyword_match.group(1).strip()
        elif competitor_match:
            role = f"竞对{int(competitor_match.group(1))}"
            part = competitor_match.group(2).upper() if competitor_match.group(2).upper() == "ASIN" else competitor_match.group(2)
            competitor_columns[col] = (role, part)
        elif col in keyword_area_columns and header not in recognized:
            keyword_columns[col] = header.strip()
        elif header not in recognized:
            unknown_headers.append(header)
    if unknown_headers:
        errors.append("存在脚本不识别的列：" + "、".join(unknown_headers))

    normalized_keyword_columns: dict[str, int] = {}
    for col, keyword_name in keyword_columns.items():
        normalized_name = normalize_name(keyword_name)
        previous_col = normalized_keyword_columns.get(normalized_name)
        if previous_col:
            errors.append(
                f"关键词表头重复：{keyword_name}（第 {previous_col} 列和第 {col} 列，忽略大小写及多余空格后相同）"
            )
        else:
            normalized_keyword_columns[normalized_name] = col

    screenshot_columns = {
        col: STATIC_FIELDS[header].field
        for col, header in headers.items()
        if header in STATIC_FIELDS and STATIC_FIELDS[header].field in SCREENSHOT_FIELD_NAMES
    }
    competitor_image_columns = {
        col: role
        for col, (role, part) in competitor_columns.items()
        if part == "操作分析"
    }
    embedded_images = extract_embedded_images(path, ws.title)
    valid_image_cells: dict[tuple[int, int], list[EmbeddedImage]] = {}
    for (row_idx, col_idx), images in embedded_images.items():
        cell_name = f"{excel_column_name(col_idx)}{row_idx}"
        if row_idx <= header_row or (
            col_idx not in screenshot_columns and col_idx not in competitor_image_columns
        ):
            errors.append(f"图片 {cell_name} 不在明细截图列中，请移动到对应日期的截图单元格")
            continue
        accepted: list[EmbeddedImage] = []
        for image in images:
            extension = Path(image.filename).suffix.lower()
            if image.content_type not in IMAGE_CONTENT_TYPES.values() or extension not in IMAGE_CONTENT_TYPES:
                errors.append(
                    f"图片 {cell_name} 格式不支持：{image.content_type}；仅支持 PNG、JPEG、GIF、WebP"
                )
                continue
            if not image.data:
                errors.append(f"图片 {cell_name} 内容为空")
                continue
            accepted.append(image)
        if accepted:
            valid_image_cells[(row_idx, col_idx)] = accepted

    active_row_indexes: set[int] = set()
    business_columns = [col for col, header in headers.items() if header in STATIC_FIELDS]
    business_columns += list(keyword_columns)
    business_columns += list(competitor_columns)
    business_columns += list(dynamic_keyword_slots)
    business_columns += list(keyword_area_columns)
    last_row = max([ws.max_row, *(row for row, _col in valid_image_cells)], default=ws.max_row)
    for row_idx in range(header_row + 1, last_row + 1):
        if any(not is_blank(ws.cell(row_idx, col).value) for col in business_columns):
            active_row_indexes.add(row_idx)
    active_row_indexes.update(row for row, _col in valid_image_cells)

    if not active_row_indexes:
        workbook.close()
        values_workbook.close()
        return [], errors

    if not country:
        errors.append("顶部配置：国家为空")
    if not asin:
        errors.append("顶部配置：ASIN 为空")

    start_day: date | None = None
    end_day: date | None = None
    try:
        start_day = parse_date(config["start_date"], workbook.epoch)
    except Exception as exc:
        errors.append(f"顶部配置：导入数据起始日期错误：{exc}")
    try:
        end_day = parse_date(config["end_date"], workbook.epoch)
    except Exception as exc:
        errors.append(f"顶部配置：导入数据终止日期错误：{exc}")
    if start_day and end_day and start_day > end_day:
        errors.append("顶部配置：导入数据起始日期不能晚于导入数据终止日期")

    rows: list[RowData] = []
    seen_keys: dict[str, int] = {}
    for row_idx in sorted(active_row_indexes):
        row_values = {header: ws.cell(row_idx, col).value for col, header in headers.items()}
        if not country or not asin or not start_day or not end_day or start_day > end_day:
            continue
        raw_day = row_values.get("日期")
        try:
            if is_blank(raw_day) or (isinstance(raw_day, str) and raw_day.lstrip().startswith("=")):
                day = start_day + timedelta(days=row_idx - header_row - 1)
            else:
                day = parse_date(raw_day, workbook.epoch)
        except Exception as exc:
            errors.append(f"第 {row_idx} 行：日期错误：{exc}")
            continue
        if day < start_day or day > end_day:
            errors.append(
                f"第 {row_idx} 行：日期 {day.isoformat()} 不在导入范围 "
                f"{start_day.isoformat()} 至 {end_day.isoformat()} 内"
            )
            continue

        country_asin, asin_country, dated_key, week_key = build_keys(country, asin, day)

        if dated_key in seen_keys:
            errors.append(f"第 {row_idx} 行：主键 {dated_key} 与第 {seen_keys[dated_key]} 行重复")
            continue
        seen_keys[dated_key] = row_idx
        parsed = RowData(row_idx, country, asin, day, country_asin, asin_country, dated_key, week_key)

        for col in keyword_area_columns:
            if col not in headers and not is_blank(ws.cell(row_idx, col).value):
                errors.append(
                    f"第 {row_idx} 行：关键词列 {excel_column_name(col)} 的第 {header_row} 行表头为空，请填写关键词名称"
                )

        field_sources: dict[tuple[str, str], tuple[str, Any]] = {}
        for header, spec in STATIC_FIELDS.items():
            if header not in row_values:
                continue
            raw_value = row_values[header]
            if is_dispimg_formula(raw_value):
                continue
            if isinstance(raw_value, str) and raw_value.lstrip().startswith("="):
                raw_value = values_ws.cell(row_idx, seen_headers[header]).value
                if is_blank(raw_value):
                    errors.append(
                        f"第 {row_idx} 行“{header}”：公式没有已保存的计算结果，请在 Excel/WPS 重新计算并保存"
                    )
                    continue
                if is_formula_error(raw_value):
                    errors.append(f"第 {row_idx} 行“{header}”：公式计算结果为 {raw_value}")
                    continue
                if isinstance(raw_value, str) and raw_value.lstrip().startswith("="):
                    errors.append(f"第 {row_idx} 行“{header}”：公式计算结果无效")
                    continue
            if is_blank(raw_value) or (
                spec.value_type in NUMERIC_VALUE_TYPES and is_numeric_placeholder(raw_value)
            ):
                continue
            try:
                converted = convert_value(raw_value, spec.value_type)
            except Exception as exc:
                errors.append(f"第 {row_idx} 行“{header}”：{exc}")
                continue
            field_key = (spec.resource, spec.field)
            max_length = TEXT_FIELD_MAX_LENGTHS.get(field_key)
            if max_length and isinstance(converted, str) and len(converted) > max_length:
                errors.append(
                    f"第 {row_idx} 行“{header}”：文本长度 {len(converted)} 超过字段上限 {max_length}"
                )
                continue
            previous = field_sources.get(field_key)
            if previous and not values_equivalent(previous[1], converted):
                errors.append(
                    f"第 {row_idx} 行“{previous[0]}”与“{header}”映射到同一字段但值不一致"
                )
                continue
            field_sources[field_key] = (header, converted)
            parsed.static.setdefault(spec.resource, {})[spec.field] = converted

        weekly_fields = parsed.static.get("weekly_performance")
        if weekly_fields:
            weekly_fields.update(calculate_weekly_derived_fields(weekly_fields))

        for col, keyword_name in keyword_columns.items():
            value = ws.cell(row_idx, col).value
            if is_blank(value) or is_dispimg_formula(value):
                continue
            if normalize_name(keyword_name) in {"xxx", "关键词名称", "示例关键词"}:
                errors.append(f"第 {row_idx} 行：词列表头仍是占位词“{keyword_name}”，请改成真实关键词")
                continue
            if is_keyword_clear_marker(value):
                parsed.keyword_clears.append(keyword_name.strip())
                continue
            parsed.keywords.append((keyword_name.strip(), str(value).strip()))

        for col, slot in dynamic_keyword_slots.items():
            if slot <= len(template_keywords) or is_blank(ws.cell(row_idx, col).value):
                continue
            errors.append(
                f"第 {row_idx} 行：第 {slot} 个关键词列没有匹配到国家 {country}、ASIN {asin} 的关键词"
            )

        competitor_parts: dict[str, dict[str, Any]] = {}
        for col, (role, part) in competitor_columns.items():
            value = ws.cell(row_idx, col).value
            if is_blank(value) or is_dispimg_formula(value):
                continue
            target = competitor_parts.setdefault(role, {})
            top_asin = top_competitor_asins.get(role)
            if top_asin:
                target["asin"] = top_asin
                target["match_by_asin"] = True
            if part == "ASIN":
                row_asin = normalize_asin(value)
                if top_asin and row_asin != top_asin:
                    errors.append(
                        f"第 {row_idx} 行：{role}明细 ASIN {row_asin} 与顶部 ASIN {top_asin} 不一致"
                    )
                    continue
                target["asin"] = row_asin
            elif part == "排名":
                target["rank"] = str(value).strip()
            else:
                target["notes"] = str(value).strip()

        for (image_row, image_col), images in valid_image_cells.items():
            role = competitor_image_columns.get(image_col)
            if image_row != row_idx or not role:
                continue
            target = competitor_parts.setdefault(role, {})
            top_asin = top_competitor_asins.get(role)
            if top_asin:
                target["asin"] = top_asin
                target["match_by_asin"] = True
            parsed.competitor_images.setdefault(role, []).extend(images)

        for role, item in competitor_parts.items():
            if (item.get("rank") or item.get("notes") or parsed.competitor_images.get(role)) and not item.get("asin"):
                errors.append(f"第 {row_idx} 行：{role}填写了日数据，但顶部和明细均未填写{role} ASIN")
            elif item.get("asin"):
                parsed.competitors[role] = item

        for (image_row, image_col), images in valid_image_cells.items():
            if image_row == row_idx and image_col in screenshot_columns:
                parsed.images.setdefault(screenshot_columns[image_col], []).extend(images)

        if (
            parsed.static
            or parsed.keywords
            or parsed.keyword_clears
            or parsed.competitors
            or parsed.images
            or parsed.competitor_images
        ):
            rows.append(parsed)

    workbook.close()
    values_workbook.close()
    return rows, errors


class NocoBaseClient:
    def __init__(self, base_url: str, token: str, role: str | None, timeout: int = 30):
        self.base_url = base_url.rstrip("/")
        self.timeout = timeout
        auth_value = token.strip()
        if not auth_value.lower().startswith("bearer "):
            auth_value = f"Bearer {auth_value}"
        self.headers = {"Authorization": auth_value, "Accept": "application/json"}
        if role:
            self.headers["X-Role"] = role

    def request(self, action: str, resource: str, *, params: dict[str, Any] | None = None,
                data: dict[str, Any] | None = None) -> dict[str, Any]:
        query = urllib.parse.urlencode(params or {})
        url = f"{self.base_url}/{resource}:{action}" + (f"?{query}" if query else "")
        body = None
        headers = dict(self.headers)
        method = "GET" if action in {"list", "get"} else "POST"
        if data is not None:
            body = json.dumps(data, ensure_ascii=False).encode("utf-8")
            headers["Content-Type"] = "application/json"
        request = urllib.request.Request(url, data=body, headers=headers, method=method)
        retry_delays = (5, 15, 30) if method == "GET" else ()
        for attempt in range(len(retry_delays) + 1):
            try:
                with urllib.request.urlopen(request, timeout=self.timeout) as response:
                    text = response.read().decode("utf-8")
                break
            except urllib.error.HTTPError as exc:
                detail = exc.read().decode("utf-8", errors="replace")
                raise ApiError(f"{method} {resource}:{action} 返回 HTTP {exc.code}: {detail[:800]}") from exc
            except urllib.error.URLError as exc:
                if attempt >= len(retry_delays):
                    raise ConnectionApiError(f"无法连接 NocoBase：{exc.reason}") from exc
                delay = retry_delays[attempt]
                print(
                    f"{resource}:{action} 连接失败，{delay} 秒后进行第 {attempt + 1}/{len(retry_delays)} 次重试。",
                    file=sys.stderr,
                )
                time.sleep(delay)
        if not text:
            return {}
        try:
            return json.loads(text)
        except json.JSONDecodeError as exc:
            raise ApiError(f"{resource}:{action} 返回了非 JSON 内容：{text[:300]}") from exc

    def list_all(self, resource: str, filter_obj: dict[str, Any], page_size: int = 200) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        page = 1
        while True:
            response = self.request("list", resource, params={
                "filter": json.dumps(filter_obj, ensure_ascii=False, separators=(",", ":")),
                "page": page,
                "pageSize": page_size,
            })
            batch = response.get("data") or []
            if not isinstance(batch, list):
                raise ApiError(f"{resource}:list 返回结构异常")
            rows.extend(batch)
            meta = response.get("meta") or {}
            total_page = int(meta.get("totalPage") or 0)
            if not batch or (total_page and page >= total_page) or len(batch) < page_size:
                break
            page += 1
        return rows

    def fetch_by_values(self, resource: str, field_name: str, values: Iterable[str | int]) -> list[dict[str, Any]]:
        unique_values = list(dict.fromkeys(value for value in values if value not in (None, "")))
        result: list[dict[str, Any]] = []
        for start in range(0, len(unique_values), 100):
            chunk = unique_values[start:start + 100]
            result.extend(self.list_all(resource, {field_name: {"$in": chunk}}))
        return result

    def create(self, resource: str, payload: dict[str, Any]) -> dict[str, Any]:
        response = self.request("create", resource, data=payload)
        data = response.get("data") or {}
        return data if isinstance(data, dict) else {}

    def update(self, resource: str, key: str | int, payload: dict[str, Any]) -> None:
        self.request("update", resource, params={"filterByTk": key}, data=payload)

    def create_static_with_recovery(
        self, resource: str, key_field: str, key: str, payload: dict[str, Any]
    ) -> dict[str, Any]:
        retry_delays = (5, 15, 30)
        for attempt in range(len(retry_delays) + 1):
            try:
                return self.create(resource, payload)
            except ConnectionApiError:
                if attempt < len(retry_delays):
                    delay = retry_delays[attempt]
                    print(
                        f"{resource}:create 主键 {key} 连接失败，{delay} 秒后回查写入结果。",
                        file=sys.stderr,
                    )
                    time.sleep(delay)
                records = self.fetch_by_values(resource, key_field, [key])
                if len(records) > 1:
                    raise ApiError(f"{resource} 主键回查返回重复记录：{key}")
                if records:
                    print(f"{resource}:create 主键 {key} 已确认写入，继续处理。")
                    return records[0]
                if attempt >= len(retry_delays):
                    raise
                print(
                    f"{resource}:create 主键 {key} 尚未写入，进行第 {attempt + 1}/{len(retry_delays)} 次重试。",
                    file=sys.stderr,
                )
        raise AssertionError("unreachable")

    def update_static_with_retry(
        self, resource: str, key: str, payload: dict[str, Any]
    ) -> None:
        retry_delays = (5, 15, 30)
        for attempt in range(len(retry_delays) + 1):
            try:
                self.update(resource, key, payload)
                return
            except ConnectionApiError:
                if attempt >= len(retry_delays):
                    raise
                delay = retry_delays[attempt]
                print(
                    f"{resource}:update 主键 {key} 连接失败，{delay} 秒后进行第 {attempt + 1}/{len(retry_delays)} 次重试。",
                    file=sys.stderr,
                )
                time.sleep(delay)

    def upload_attachment(self, image: EmbeddedImage) -> str:
        boundary = f"----daily-history-{uuid.uuid4().hex}"
        safe_filename = re.sub(r"[^A-Za-z0-9._-]", "_", image.filename)
        prefix = (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; filename="{safe_filename}"\r\n'
            f"Content-Type: {image.content_type}\r\n\r\n"
        ).encode("ascii")
        body = prefix + image.data + f"\r\n--{boundary}--\r\n".encode("ascii")
        headers = dict(self.headers)
        headers["Content-Type"] = f"multipart/form-data; boundary={boundary}"
        headers["Content-Length"] = str(len(body))
        request = urllib.request.Request(
            f"{self.base_url}/attachments:upload",
            data=body,
            headers=headers,
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=self.timeout) as response:
                text = response.read().decode("utf-8")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise ApiError(f"上传图片 {image.filename} 返回 HTTP {exc.code}: {detail[:800]}") from exc
        except urllib.error.URLError as exc:
            raise ApiError(f"上传图片 {image.filename} 失败：{exc.reason}") from exc
        try:
            payload = json.loads(text)
        except json.JSONDecodeError as exc:
            raise ApiError(f"上传图片 {image.filename} 返回了非 JSON 内容：{text[:300]}") from exc
        data = payload.get("data") if isinstance(payload, dict) else None
        url = data.get("url") if isinstance(data, dict) else None
        if not url and isinstance(payload, dict):
            url = payload.get("url")
        if not url:
            raise ApiError(f"上传图片 {image.filename} 后未返回 url")
        return str(url)


def unique_by(records: Iterable[dict[str, Any]], key_field: str, label: str) -> dict[Any, dict[str, Any]]:
    result: dict[Any, dict[str, Any]] = {}
    duplicates: list[Any] = []
    for record in records:
        key = record.get(key_field)
        if key in (None, ""):
            continue
        if key in result:
            duplicates.append(key)
        else:
            result[key] = record
    if duplicates:
        preview = "、".join(str(value) for value in list(dict.fromkeys(duplicates))[:10])
        raise ValidationError(f"线上{label}存在重复定位值：{preview}")
    return result


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")


@dataclass
class OnlineState:
    static: dict[str, dict[Any, dict[str, Any]]]
    asin_master: dict[str, dict[str, Any]]
    keywords: dict[tuple[str, str], dict[str, Any]]
    keyword_positions: dict[tuple[int, str], dict[str, Any]]
    competitors: dict[tuple[str, str], dict[str, Any]]
    competitors_by_asin: dict[tuple[str, str], list[dict[str, Any]]]
    competitor_daily: dict[tuple[int, str], dict[str, Any]]


def load_online_state(client: NocoBaseClient, rows: list[RowData]) -> OnlineState:
    dated_keys = [row.country_asin_date for row in rows]
    country_asin_keys = [row.country_asin for row in rows]
    asin_country_keys = [row.asin_country for row in rows]
    static: dict[str, dict[Any, dict[str, Any]]] = {}
    for resource, key_field in RESOURCE_KEY_FIELDS.items():
        records = client.fetch_by_values(resource, key_field, dated_keys)
        static[resource] = unique_by(records, key_field, resource)

    asin_master_records = client.fetch_by_values("asin", "unique", asin_country_keys)
    asin_master = unique_by(asin_master_records, "unique", "ASIN主数据")

    keyword_records = client.fetch_by_values("sqp_keywords", "country_asin", country_asin_keys)
    keywords: dict[tuple[str, str], dict[str, Any]] = {}
    for record in keyword_records:
        key = (str(record.get("country_asin") or ""), normalize_name(record.get("keyword_name")))
        if not key[0] or not key[1]:
            continue
        if key in keywords:
            raise ValidationError(f"线上关键词重复：{key[0]} / {record.get('keyword_name')}")
        keywords[key] = record

    keyword_position_records = client.fetch_by_values(
        "sqp_keyword_daily_positions", "country_asin_date", dated_keys
    )
    keyword_positions: dict[tuple[int, str], dict[str, Any]] = {}
    for record in keyword_position_records:
        keyword_id = record.get("sqp_keyword_id")
        dated_key = str(record.get("country_asin_date") or "")
        if keyword_id in (None, "") or not dated_key:
            continue
        key = (int(keyword_id), dated_key)
        if key in keyword_positions:
            raise ValidationError(f"线上关键词每日自然位重复：关键词ID {keyword_id} / {dated_key}")
        keyword_positions[key] = record

    competitor_records = client.fetch_by_values("order_link_competitor_asins", "country_asin", country_asin_keys)
    competitors: dict[tuple[str, str], dict[str, Any]] = {}
    competitors_by_asin: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for record in competitor_records:
        country_asin = str(record.get("country_asin") or "")
        role = str(record.get("role") or "").strip()
        competitor_asin = normalize_asin(record.get("competitor_asin"))
        key = (country_asin, role)
        if not country_asin or not role:
            continue
        if key in competitors:
            raise ValidationError(f"线上竞对位重复：{key[0]} / {key[1]}")
        competitors[key] = record
        if competitor_asin:
            competitors_by_asin.setdefault((country_asin, competitor_asin), []).append(record)

    competitor_daily_records = client.fetch_by_values(
        "order_link_competitor_asins_daily", "country_asin_date", dated_keys
    )
    competitor_daily: dict[tuple[int, str], dict[str, Any]] = {}
    for record in competitor_daily_records:
        competitor_id = record.get("competitor_id")
        dated_key = str(record.get("country_asin_date") or "")
        if competitor_id in (None, "") or not dated_key:
            continue
        key = (int(competitor_id), dated_key)
        if key in competitor_daily:
            raise ValidationError(f"线上竞对每日记录重复：竞对ID {competitor_id} / {dated_key}")
        competitor_daily[key] = record

    return OnlineState(
        static,
        asin_master,
        keywords,
        keyword_positions,
        competitors,
        competitors_by_asin,
        competitor_daily,
    )


def competitor_match_key(row: RowData, role: str, item: dict[str, Any]) -> tuple[str, str, str]:
    if item.get("match_by_asin"):
        return "asin", row.country_asin, item["asin"]
    return "role", row.country_asin, role


def find_existing_competitor(
    row: RowData, role: str, item: dict[str, Any], state: OnlineState
) -> dict[str, Any] | None:
    if not item.get("match_by_asin"):
        return state.competitors.get((row.country_asin, role))
    matches = state.competitors_by_asin.get((row.country_asin, item["asin"]), [])
    if len(matches) > 1:
        roles = "、".join(str(record.get("role") or "") for record in matches)
        raise ValidationError(
            f"线上 {row.country_asin} 的竞对 ASIN {item['asin']} 重复绑定到多个竞对位：{roles}"
        )
    return matches[0] if matches else None


def next_available_competitor_role(country_asin: str, state: OnlineState) -> str:
    used_numbers: set[int] = set()
    for existing_country_asin, role in state.competitors:
        if existing_country_asin != country_asin:
            continue
        match = re.fullmatch(r"竞对([1-9]\d*)", role)
        if match:
            used_numbers.add(int(match.group(1)))
    position = 1
    while position in used_numbers:
        position += 1
    return f"竞对{position}"


def validate_competitor_consistency(rows: list[RowData], state: OnlineState) -> None:
    requested: dict[tuple[str, str], tuple[str, int]] = {}
    for row in rows:
        for role, item in row.competitors.items():
            if item.get("match_by_asin"):
                find_existing_competitor(row, role, item, state)
                continue
            key = (row.country_asin, role)
            asin = item["asin"]
            previous = requested.get(key)
            if previous and previous[0] != asin:
                raise ValidationError(
                    f"第 {row.excel_row} 行：{row.country_asin} 的 {role} ASIN 与第 {previous[1]} 行不一致"
                )
            requested[key] = (asin, row.excel_row)
            existing = state.competitors.get(key)
            if existing and normalize_asin(existing.get("competitor_asin")) != asin:
                raise ValidationError(
                    f"第 {row.excel_row} 行：线上 {row.country_asin} / {role} 已绑定 "
                    f"{existing.get('competitor_asin')}，模板填写为 {asin}"
                )


def row_resource_fields(row: RowData) -> dict[str, dict[str, Any]]:
    resources = {resource: fields for resource, fields in row.static.items()}
    if row.images:
        resources.setdefault("daily_order_link_tracking", {})
    return resources


def build_summary(rows: list[RowData], state: OnlineState) -> dict[str, dict[str, int]]:
    summary: dict[str, dict[str, int]] = {
        resource: {"create": 0, "update": 0} for resource in RESOURCE_KEY_FIELDS
    }
    summary["sqp_keywords"] = {"create": 0, "reuse": 0}
    summary["sqp_keyword_daily_positions"] = {"create": 0, "update": 0, "clear": 0}
    summary["order_link_competitor_asins"] = {"create": 0, "reuse": 0}
    summary["order_link_competitor_asins_daily"] = {"create": 0, "update": 0}
    summary["attachments"] = {
        "upload": sum(
            len(images)
            for row in rows
            for image_groups in (row.images, row.competitor_images)
            for images in image_groups.values()
        )
    }
    counted_keywords: set[tuple[str, str]] = set()
    counted_competitors: set[tuple[str, str, str]] = set()

    for row in rows:
        for resource, fields in row_resource_fields(row).items():
            if not fields and not (resource == "daily_order_link_tracking" and row.images):
                continue
            key = row.country_asin_week if resource == "weekly_performance" else row.country_asin_date
            action = "update" if key in state.static[resource] else "create"
            summary[resource][action] += 1
        for keyword_name, _rank in row.keywords:
            keyword_key = (row.country_asin, normalize_name(keyword_name))
            existing = state.keywords.get(keyword_key)
            if keyword_key not in counted_keywords:
                summary["sqp_keywords"]["reuse" if existing else "create"] += 1
                counted_keywords.add(keyword_key)
            if existing:
                daily_key = (int(existing["id"]), row.country_asin_date)
                action = "update" if daily_key in state.keyword_positions else "create"
            else:
                action = "create"
            summary["sqp_keyword_daily_positions"][action] += 1
        for keyword_name in row.keyword_clears:
            keyword_key = (row.country_asin, normalize_name(keyword_name))
            existing_keyword = state.keywords.get(keyword_key)
            if not existing_keyword:
                continue
            daily_key = (int(existing_keyword["id"]), row.country_asin_date)
            existing_position = state.keyword_positions.get(daily_key)
            if existing_position and not is_blank(existing_position.get("actual_rank")):
                summary["sqp_keyword_daily_positions"]["clear"] += 1
        for role, item in row.competitors.items():
            competitor_key = competitor_match_key(row, role, item)
            existing = find_existing_competitor(row, role, item, state)
            if competitor_key not in counted_competitors:
                summary["order_link_competitor_asins"]["reuse" if existing else "create"] += 1
                counted_competitors.add(competitor_key)
            if existing:
                daily_key = (int(existing["id"]), row.country_asin_date)
                action = "update" if daily_key in state.competitor_daily else "create"
            else:
                action = "create"
            summary["order_link_competitor_asins_daily"][action] += 1
    return summary


def print_summary(summary: dict[str, dict[str, int]], title: str) -> None:
    print(f"\n{title}")
    print("-" * len(title))
    for resource, counts in summary.items():
        detail = "，".join(f"{name}={count}" for name, count in counts.items() if count)
        if detail:
            print(f"{resource}: {detail}")


def static_create_payload(resource: str, row: RowData, fields: dict[str, Any], state: OnlineState) -> dict[str, Any]:
    payload = dict(fields)
    if resource == "daily_asins":
        payload.update({
            "country_asin_date": row.country_asin_date,
            "asin_country": row.asin_country,
            "country": row.country,
            "asin": row.asin,
            "date": row.day.isoformat(),
        })
        master = state.asin_master.get(row.asin_country) or {}
        for field_name in ("model", "sale_owner", "model_sales"):
            if not is_blank(master.get(field_name)):
                payload[field_name] = master[field_name]
        if "selling_accounts" not in payload and not is_blank(master.get("selling_accounts")):
            payload["selling_accounts"] = master["selling_accounts"]
    elif resource == "weekly_performance":
        payload.update({
            "country_asin_week": row.country_asin_week,
            "country": row.country,
            "asin": row.asin,
            "date": row.day.isoformat(),
        })
    else:
        payload[RESOURCE_KEY_FIELDS[resource]] = row.country_asin_date
    return payload


def materialize_screenshot_fields(
    client: NocoBaseClient, row: RowData, fields: dict[str, Any]
) -> tuple[dict[str, Any], int]:
    materialized = dict(fields)
    uploaded = 0
    for field_name, images in row.images.items():
        parts: list[str] = []
        current = str(materialized.get(field_name) or "").strip()
        if current:
            parts.append(current)
        for image in images:
            url = client.upload_attachment(image)
            parts.append(f"![截图]({url})")
            uploaded += 1
        materialized[field_name] = "\n\n".join(parts)
    return materialized, uploaded


def materialize_images(
    client: NocoBaseClient, value: Any, images: list[EmbeddedImage]
) -> tuple[str, int]:
    parts: list[str] = []
    current = str(value or "").strip()
    if current:
        parts.append(current)
    for image in images:
        url = client.upload_attachment(image)
        parts.append(f"![截图]({url})")
    return "\n\n".join(parts), len(images)


def execute_import(client: NocoBaseClient, rows: list[RowData], state: OnlineState) -> dict[str, dict[str, int]]:
    result: dict[str, dict[str, int]] = {
        resource: {"created": 0, "updated": 0} for resource in RESOURCE_KEY_FIELDS
    }
    result["sqp_keywords"] = {"created": 0, "reused": 0}
    result["sqp_keyword_daily_positions"] = {"created": 0, "updated": 0, "cleared": 0}
    result["order_link_competitor_asins"] = {"created": 0, "reused": 0}
    result["order_link_competitor_asins_daily"] = {"created": 0, "updated": 0}
    result["attachments"] = {"uploaded": 0}

    for row in rows:
        for resource, original_fields in row_resource_fields(row).items():
            fields = dict(original_fields)
            if resource == "daily_order_link_tracking" and row.images:
                fields, uploaded = materialize_screenshot_fields(client, row, fields)
                result["attachments"]["uploaded"] += uploaded
            if not fields:
                continue
            key = row.country_asin_week if resource == "weekly_performance" else row.country_asin_date
            if key in state.static[resource]:
                client.update_static_with_retry(resource, key, fields)
                result[resource]["updated"] += 1
            else:
                created = client.create_static_with_recovery(
                    resource,
                    RESOURCE_KEY_FIELDS[resource],
                    key,
                    static_create_payload(resource, row, fields, state),
                )
                state.static[resource][key] = created or {RESOURCE_KEY_FIELDS[resource]: key}
                result[resource]["created"] += 1

    reused_keywords: set[tuple[str, str]] = set()
    for row in rows:
        for keyword_name, rank in row.keywords:
            keyword_key = (row.country_asin, normalize_name(keyword_name))
            keyword = state.keywords.get(keyword_key)
            if not keyword:
                now = utc_timestamp()
                keyword = client.create("sqp_keywords", {
                    "country_asin": row.country_asin,
                    "country": row.country,
                    "asin": row.asin,
                    "keyword_name": keyword_name,
                    "created_at": now,
                    "updated_at": now,
                })
                if not keyword.get("id"):
                    raise ApiError(f"创建关键词后未返回 id：{row.country_asin} / {keyword_name}")
                state.keywords[keyword_key] = keyword
                result["sqp_keywords"]["created"] += 1
            elif keyword_key not in reused_keywords:
                result["sqp_keywords"]["reused"] += 1
            reused_keywords.add(keyword_key)
            keyword_id = int(keyword["id"])
            daily_key = (keyword_id, row.country_asin_date)
            existing = state.keyword_positions.get(daily_key)
            if existing:
                client.update("sqp_keyword_daily_positions", existing["id"], {
                    "actual_rank": rank,
                    "updated_at": utc_timestamp(),
                })
                result["sqp_keyword_daily_positions"]["updated"] += 1
            else:
                now = utc_timestamp()
                created = client.create("sqp_keyword_daily_positions", {
                    "country_asin_date": row.country_asin_date,
                    "country_asin": row.country_asin,
                    "country": row.country,
                    "asin": row.asin,
                    "sqp_keyword_id": keyword_id,
                    "date": row.day.isoformat(),
                    "actual_rank": rank,
                    "created_at": now,
                    "updated_at": now,
                })
                state.keyword_positions[daily_key] = created
                result["sqp_keyword_daily_positions"]["created"] += 1

    for row in rows:
        for keyword_name in row.keyword_clears:
            keyword_key = (row.country_asin, normalize_name(keyword_name))
            keyword = state.keywords.get(keyword_key)
            if not keyword:
                continue
            daily_key = (int(keyword["id"]), row.country_asin_date)
            existing = state.keyword_positions.get(daily_key)
            if not existing or is_blank(existing.get("actual_rank")):
                continue
            client.update("sqp_keyword_daily_positions", existing["id"], {
                "actual_rank": None,
                "updated_at": utc_timestamp(),
            })
            existing["actual_rank"] = None
            result["sqp_keyword_daily_positions"]["cleared"] += 1

    reused_competitors: set[tuple[str, str, str]] = set()
    for row in rows:
        for role, item in row.competitors.items():
            competitor_key = competitor_match_key(row, role, item)
            competitor = find_existing_competitor(row, role, item, state)
            if not competitor:
                assigned_role = (
                    next_available_competitor_role(row.country_asin, state)
                    if item.get("match_by_asin")
                    else role
                )
                competitor = client.create("order_link_competitor_asins", {
                    "country_asin": row.country_asin,
                    "competitor_asin": item["asin"],
                    "role": assigned_role,
                })
                if not competitor.get("id"):
                    raise ApiError(f"创建竞对后未返回 id：{row.country_asin} / {assigned_role}")
                state.competitors[(row.country_asin, assigned_role)] = competitor
                state.competitors_by_asin.setdefault(
                    (row.country_asin, item["asin"]), []
                ).append(competitor)
                result["order_link_competitor_asins"]["created"] += 1
            elif competitor_key not in reused_competitors:
                result["order_link_competitor_asins"]["reused"] += 1
            reused_competitors.add(competitor_key)
            competitor_id = int(competitor["id"])
            daily_key = (competitor_id, row.country_asin_date)
            daily_fields = {name: item[name] for name in ("rank", "notes") if item.get(name) is not None}
            competitor_images = row.competitor_images.get(role, [])
            if competitor_images:
                notes, uploaded = materialize_images(client, item.get("notes"), competitor_images)
                daily_fields["notes"] = notes
                result["attachments"]["uploaded"] += uploaded
            if not daily_fields:
                continue
            existing = state.competitor_daily.get(daily_key)
            if existing:
                client.update("order_link_competitor_asins_daily", existing["id"], daily_fields)
                result["order_link_competitor_asins_daily"]["updated"] += 1
            else:
                created = client.create("order_link_competitor_asins_daily", {
                    "country_asin_date": row.country_asin_date,
                    "competitor_id": competitor_id,
                    "date": row.day.isoformat(),
                    **daily_fields,
                })
                state.competitor_daily[daily_key] = created
                result["order_link_competitor_asins_daily"]["created"] += 1
    return result


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="NocoBase 每日历史数据 Excel 更新/插入工具")
    input_group = parser.add_mutually_exclusive_group()
    input_group.add_argument(
        "--file",
        type=Path,
        help=f"待导入的 .xlsx 文件；不指定 --file/--folder 时默认：{DEFAULT_EXCEL_FILE}",
    )
    input_group.add_argument(
        "--folder",
        type=Path,
        help="批量导入文件夹；只处理第一层 .xlsx 文件，忽略 ~$ 开头的 Excel 临时文件",
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help=f"数据工作表，默认：{DEFAULT_SHEET}")
    parser.add_argument("--api-base-url", default=os.getenv("NOCOBASE_API_URL", DEFAULT_API_BASE_URL))
    parser.add_argument(
        "--token",
        default=DEFAULT_API_TOKEN or os.getenv("NOCOBASE_API_TOKEN"),
        help="优先读取代码内 DEFAULT_API_TOKEN，也可通过 --token 覆盖；代码内为空时读取 NOCOBASE_API_TOKEN",
    )
    parser.add_argument("--role", default=os.getenv("NOCOBASE_ROLE"), help="可选 X-Role")
    parser.add_argument("--timeout", type=int, default=30)
    parser.add_argument("--validate-only", action="store_true", help="只校验 Excel，不连接 NocoBase")
    parser.add_argument("--execute", action="store_true", help="执行写入；不指定时默认只做线上 dry-run")
    parser.add_argument("--confirm-production", default="", help=f"生产写入确认词：{EXECUTE_CONFIRMATION}")
    return parser.parse_args(argv)


def resolve_input_files(args: argparse.Namespace) -> list[Path]:
    if args.folder is None:
        file_path = args.file or DEFAULT_EXCEL_FILE
        if not file_path.exists():
            raise ValidationError(f"文件不存在：{file_path}")
        if not file_path.is_file():
            raise ValidationError(f"不是文件：{file_path}")
        if file_path.suffix.lower() != ".xlsx":
            raise ValidationError("只支持 .xlsx 文件")
        return [file_path]

    if not args.folder.exists():
        raise ValidationError(f"文件夹不存在：{args.folder}")
    if not args.folder.is_dir():
        raise ValidationError(f"不是文件夹：{args.folder}")
    try:
        files = sorted(
            (
                path
                for path in args.folder.iterdir()
                if path.is_file()
                and path.suffix.lower() == ".xlsx"
                and not path.name.startswith("~$")
            ),
            key=lambda path: path.name.casefold(),
        )
    except OSError as exc:
        raise ValidationError(f"无法读取文件夹 {args.folder}：{exc}") from exc
    if not files:
        raise ValidationError(f"文件夹中没有可处理的 .xlsx 文件：{args.folder}")
    return files


def process_file(file_path: Path, args: argparse.Namespace) -> int:

    try:
        rows, errors = read_rows(file_path, args.sheet)
        if errors:
            print("Excel 校验失败：", file=sys.stderr)
            for error in errors:
                print(f"- {error}", file=sys.stderr)
            return 2
        print(f"Excel 校验通过：有效数据 {len(rows)} 行")
        image_count = sum(
            len(images)
            for row in rows
            for image_groups in (row.images, row.competitor_images)
            for images in image_groups.values()
        )
        if image_count:
            print(f"识别到嵌入图片 {image_count} 张；仅在正式执行时上传。")
        if not rows:
            print("没有可导入的业务数据；模板中的空白预留行不会被处理。")
            return 0
        if args.validate_only:
            print("validate-only 完成：未连接 NocoBase，未写入任何数据。")
            return 0
        if not args.token:
            print(
                "缺少 API Token：请在代码顶部填写 DEFAULT_API_TOKEN，或设置 NOCOBASE_API_TOKEN / 传入 --token",
                file=sys.stderr,
            )
            return 2

        client = NocoBaseClient(args.api_base_url, args.token, args.role, args.timeout)
        state = load_online_state(client, rows)
        validate_competitor_consistency(rows, state)
        summary = build_summary(rows, state)
        print_summary(summary, "线上预检结果")
        if not args.execute:
            print("\nDry-run 完成：只读取线上数据，没有上传附件，也没有执行创建或更新。")
            return 0

        print("\n开始执行生产写入。发生首个 API 错误时脚本会立即停止。")
        result = execute_import(client, rows, state)
        print_summary(result, "生产写入结果")
        return 0
    except (ValidationError, ApiError) as exc:
        print(f"执行失败：{exc}", file=sys.stderr)
        return 2
    except KeyboardInterrupt:
        print("用户中断。", file=sys.stderr)
        return 130


def process_batch(files: list[Path], args: argparse.Namespace) -> int:
    for index, file_path in enumerate(files, start=1):
        print(f"\n===== [{index}/{len(files)}] {file_path.name} =====")
        result = process_file(file_path, args)
        if result != 0:
            print(
                f"批量处理停止：{file_path.name} 失败；此前成功完成 {index - 1} 个文件。",
                file=sys.stderr,
            )
            return result
    return 0


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.execute and args.confirm_production != EXECUTE_CONFIRMATION:
        print(
            f"拒绝写入：--execute 必须同时提供 --confirm-production {EXECUTE_CONFIRMATION}",
            file=sys.stderr,
        )
        return 2

    try:
        files = resolve_input_files(args)
    except ValidationError as exc:
        print(f"执行失败：{exc}", file=sys.stderr)
        return 2

    if args.folder is None:
        return process_file(files[0], args)

    if args.execute:
        preflight_args = argparse.Namespace(**vars(args))
        preflight_args.execute = False
        print("\n===== 批量生产导入前检查：所有文件只读预检 =====")
        result = process_batch(files, preflight_args)
        if result != 0:
            print("生产写入尚未开始，本次没有创建、更新或上传任何数据。", file=sys.stderr)
            return result
        print(f"\n批量生产导入前检查完成：{len(files)} 个文件全部通过。")
        print("\n===== 开始批量生产写入 =====")

    result = process_batch(files, args)
    if result != 0:
        return result

    if args.validate_only:
        mode = "批量校验"
    elif args.execute:
        mode = "批量生产导入"
    else:
        mode = "批量线上预检"
    print(f"\n{mode}完成：成功处理 {len(files)} 个文件。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
